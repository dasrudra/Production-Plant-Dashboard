from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_SHEET_NAME = "Plan-KPP"


def safe_number(value: Any) -> float:
    """
    Convert Excel values into a usable number.

    Examples:
    500000 -> 500000
    "500,000" -> 500000
    "-" -> 0
    None -> 0
    """
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if text in ["", "-", "—", "N/A", "n/a", "None"]:
        return 0

    text = text.replace(",", "")

    try:
        return float(text)
    except ValueError:
        return 0


def round_number(value: float, digits: int = 2) -> float:
    return round(float(value), digits)


def extract_month_and_division(activity_title: str) -> dict[str, str]:
    """
    Example:
    Activity Plan for Month of July'26 | KPP Division | Complete Performance Data
    """
    result = {
        "month": "",
        "division": "",
    }

    if not activity_title:
        return result

    month_match = re.search(r"Month of\s+([^|]+)", activity_title, re.IGNORECASE)
    if month_match:
        result["month"] = month_match.group(1).strip()

    title_parts = [part.strip() for part in activity_title.split("|")]
    for part in title_parts:
        if "division" in part.lower():
            result["division"] = part

    return result


def calculate_status(achievement_decimal: float) -> str:
    """
    Status rule:
    achievement >= 90% = Excellent
    achievement >= 80% = Good
    achievement >= 60% = Normal
    below 60% = Low
    """
    if achievement_decimal >= 0.90:
        return "Excellent"

    if achievement_decimal >= 0.80:
        return "Good"

    if achievement_decimal >= 0.60:
        return "Normal"

    return "Low"


def analyze_capacity_plan_excel(file_path: str | Path) -> dict[str, Any]:
    """
    Analyze KPP production activity plan Excel file.

    Current expected source sheet:
    Plan-KPP

    Current expected table:
    Row 7 = Header row
    Row 8 onward = Machine rows
    TOTAL row = Stop point
    """
    file_path = Path(file_path)

    workbook = load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames

    if REQUIRED_SHEET_NAME not in sheet_names:
        return {
            "success": False,
            "message": "Invalid Excel template. Plan-KPP sheet was not found.",
            "fileName": file_path.name,
            "sheetNames": sheet_names,
            "requiredSheet": REQUIRED_SHEET_NAME,
        }

    sheet = workbook[REQUIRED_SHEET_NAME]

    report_title = str(sheet["A1"].value or "").strip()
    activity_title = str(sheet["A2"].value or "").strip()
    title_info = extract_month_and_division(activity_title)

    working_section = str(sheet["A5"].value or "").strip()

    machine_rows: list[dict[str, Any]] = []

    for row_index in range(8, sheet.max_row + 1):
        machine_type_value = sheet.cell(row=row_index, column=1).value

        if machine_type_value is None:
            continue

        machine_type = str(machine_type_value).strip()

        if machine_type == "":
            continue

        if machine_type.upper() == "TOTAL":
            break

        labour_working_day = safe_number(sheet.cell(row=row_index, column=2).value)
        machine_working_day = safe_number(sheet.cell(row=row_index, column=3).value)
        active_labor = safe_number(sheet.cell(row=row_index, column=4).value)
        labour_shift_hour = safe_number(sheet.cell(row=row_index, column=5).value)
        active_machine = safe_number(sheet.cell(row=row_index, column=6).value)
        machine_shift_hour = safe_number(sheet.cell(row=row_index, column=7).value)
        monthly_target = safe_number(sheet.cell(row=row_index, column=8).value)
        monthly_capacity = safe_number(sheet.cell(row=row_index, column=9).value)

        achievement_decimal = monthly_target / monthly_capacity if monthly_capacity > 0 else 0
        achievement_percent = achievement_decimal * 100

        labour_operation_hour_per_day = achievement_decimal * labour_shift_hour
        machine_operation_hour_per_day = achievement_decimal * machine_shift_hour

        labour_plan_minutes_per_month = (
            labour_working_day
            * active_labor
            * labour_operation_hour_per_day
            * 60
        )

        machine_plan_minutes_per_month = (
            machine_working_day
            * achievement_decimal
            * machine_operation_hour_per_day
            * 60
        )

        status = calculate_status(achievement_decimal)

        machine_rows.append(
            {
                "rowNumber": row_index,
                "machineType": machine_type,
                "labourWorkingDay": round_number(labour_working_day),
                "machineWorkingDay": round_number(machine_working_day),
                "activeLabor": round_number(active_labor),
                "labourShiftHour": round_number(labour_shift_hour),
                "activeMachine": round_number(active_machine),
                "machineShiftHour": round_number(machine_shift_hour),
                "monthlyTarget": round_number(monthly_target),
                "monthlyCapacity": round_number(monthly_capacity),
                "achievement": round_number(achievement_percent),
                "achievementDecimal": round_number(achievement_decimal, 4),
                "labourOperationHourPerDay": round_number(labour_operation_hour_per_day),
                "machineOperationHourPerDay": round_number(machine_operation_hour_per_day),
                "labourPlanMinutesPerMonth": round_number(labour_plan_minutes_per_month),
                "machinePlanMinutesPerMonth": round_number(machine_plan_minutes_per_month),
                "status": status,
            }
        )

    total_monthly_target = sum(row["monthlyTarget"] for row in machine_rows)
    total_monthly_capacity = sum(row["monthlyCapacity"] for row in machine_rows)
    total_active_labor = sum(row["activeLabor"] for row in machine_rows)
    total_active_machine = sum(row["activeMachine"] for row in machine_rows)
    total_labour_plan_minutes = sum(row["labourPlanMinutesPerMonth"] for row in machine_rows)
    total_machine_plan_minutes = sum(row["machinePlanMinutesPerMonth"] for row in machine_rows)

    capacity_utilization = (
        total_monthly_target / total_monthly_capacity * 100
        if total_monthly_capacity > 0
        else 0
    )

    status_counts: dict[str, int] = {}

    for row in machine_rows:
        status = row["status"]
        status_counts[status] = status_counts.get(status, 0) + 1

    summary = {
        "reportTitle": report_title,
        "activityTitle": activity_title,
        "month": title_info["month"],
        "division": title_info["division"],
        "workingSection": working_section,
        "monthlyTarget": round_number(total_monthly_target),
        "monthlyCapacity": round_number(total_monthly_capacity),
        "activeLabor": round_number(total_active_labor),
        "activeMachine": round_number(total_active_machine),
        "labourPlanMinutes": round_number(total_labour_plan_minutes),
        "machinePlanMinutes": round_number(total_machine_plan_minutes),
        "capacityUtilization": round_number(capacity_utilization),
        "machineCount": len(machine_rows),
        "statusCounts": status_counts,
    }

    return {
        "success": True,
        "message": "Excel file analyzed successfully.",
        "fileName": file_path.name,
        "workbookSheets": sheet_names,
        "sourceSheet": REQUIRED_SHEET_NAME,
        "activeSheets": [REQUIRED_SHEET_NAME],
        "excludedSheets": ["Sheet1"] if "Sheet1" in sheet_names else [],
        "referenceSheets": [
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in [REQUIRED_SHEET_NAME, "Sheet1"]
        ],
        "summary": summary,
        "machineRows": machine_rows,
    }
